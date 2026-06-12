"""
策略对日度数据 — 数据清洗脚本
================================================
读取原始 Excel（只读，不修改），清洗后输出 CSV。

使用方法：
    python clean_data.py

可调整配置：
    SOURCE_FILE  原始 Excel 路径
    OUTPUT_DIR   清洗结果输出目录
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ── 配置 ────────────────────────────────────────────────────────────
SOURCE_FILE = r"C:\VS CODE\project\策略对日度更新\策略对日度数据修复版.xlsx"
OUTPUT_DIR  = Path(__file__).parent
# ────────────────────────────────────────────────────────────────────


def _ws_to_df(ws, date_col_idx, data_slice, header_row=0, data_start_row=1):
    """
    从 openpyxl worksheet 提取时序数据。

    Parameters
    ----------
    date_col_idx : int        日期列的 0-based 列索引
    data_slice   : slice      数据列的 0-based 列索引切片
    header_row   : int        列名所在行的 0-based 行索引
    data_start_row : int      数据起始行的 0-based 行索引
    """
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    col_names = list(rows[header_row][data_slice])

    records = []
    for row in rows[data_start_row:]:
        date_val = row[date_col_idx]
        # 跳过日期为空或非 datetime 的行（Bloomberg 模板末尾的空行）
        if not hasattr(date_val, "year"):
            continue
        data_vals = [
            None if str(v) in ("#N/A", "#NAME?", "None") else v
            for v in row[data_slice]
        ]
        records.append([date_val] + data_vals)

    df = pd.DataFrame(records, columns=["Date"] + col_names)
    df["Date"] = pd.to_datetime(df["Date"]).dt.normalize()

    # 去重（保留最新记录）
    df = df.drop_duplicates(subset="Date", keep="first")
    # 升序排列
    df = df.sort_values("Date").reset_index(drop=True)

    # 数据列统一转 float
    for col in col_names:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


_OVERSEAS_RENAME = {
    "CME金":  "GC1 Comdty",
    "CME银":  "SI1 Comdty",
    "CME铜":  "HG1 Comdty",
    "LME铝":  "LMAHDS03 Comdty",
    "LME铜":  "LMCADS03 LME Comdty",
    "LME镍":  "LMNIDS03 LME Comdty",
    "LME锌":  "LMZSDS03 LME Comdty",
    "LME锡":  "LMSNDS03 LME Comdty",
}

_DOMESTIC_RENAME = {
    "铜": "CU2 Comdty",
    "铝": "AA2 Comdty",
    "金": "AUA2 Comdty",
    "银": "SAI2 COMB Comdty",
    "镍": "XII2 Comdty",
    "锌": "ZNA2 Comdty",
    "锡": "XOO2 Comdty",
}


def extract_overseas(wb):
    """外盘下午三点价格：E1:L1 列名，D 列日期，E:L 数据。"""
    ws = wb["外盘下午三点价格"]
    df = _ws_to_df(ws, date_col_idx=3, data_slice=slice(4, 12),
                   header_row=0, data_start_row=1)
    return df.rename(columns=_OVERSEAS_RENAME)


def extract_domestic(wb):
    """内盘收盘价：D1:J1 列名，C 列日期，D:J 数据。"""
    ws = wb["内盘收盘价"]
    df = _ws_to_df(ws, date_col_idx=2, data_slice=slice(3, 10),
                   header_row=0, data_start_row=1)
    return df.rename(columns=_DOMESTIC_RENAME)


def extract_cnh(wb):
    """离岸人民币：B1 列名，A 列日期，第 3 行起为数据。"""
    ws = wb["离岸人民币"]
    return _ws_to_df(ws, date_col_idx=0, data_slice=slice(1, 2),
                     header_row=0, data_start_row=2)


def add_derived_columns(df):
    """新增衍生价差列。"""
    usd = df["USDCNY REGN Curncy"]

    df["CU COMEX-LME价差"] = df["HG1 Comdty"] * 22.04 - df["LMCADS03 LME Comdty"]
    df["CU SHFE-LME价差"]  = df["CU2 Comdty"]  / usd - df["LMCADS03 LME Comdty"] * 1.13
    df["AL SHFE-LME价差"]  = df["AA2 Comdty"]  / usd - df["LMAHDS03 Comdty"]      * 1.13

    # CME 金银比
    df["CME金银比"] = df["GC1 Comdty"] / df["SI1 Comdty"]

    # 黄金跨市价差：内盘金(元/克) ÷ 汇率 − CME金(美元/盎司) ÷ 31.1035
    df["黄金SHFE-CME"] = df["AUA2 Comdty"] / usd - df["GC1 Comdty"] / 31.1035

    # 白银跨市价差：内盘银(元/千克) ÷ 1000 ÷ 汇率 − CME银(美元/盎司) ÷ 31.1035 × 1.13
    df["白银SHFE-CME"] = (
        df["SAI2 COMB Comdty"] / 1000 / usd - df["SI1 Comdty"] / 31.1035 * 1.13
    )

    # 镍/锌/锡跨市价差
    df["镍SHFE-LME"] = df["XII2 Comdty"] / usd - df["LMNIDS03 LME Comdty"] * 1.13
    df["锌SHFE-LME"] = df["ZNA2 Comdty"] / usd - df["LMZSDS03 LME Comdty"] * 1.13
    df["锡SHFE-LME"] = df["XOO2 Comdty"] / usd - df["LMSNDS03 LME Comdty"] * 1.13

    return df


def merge_all(df_overseas, df_domestic, df_cnh):
    """三个 sheet 按 Date outer join 合并，升序。"""
    df = (df_overseas
          .merge(df_domestic, on="Date", how="outer")
          .merge(df_cnh,      on="Date", how="outer"))
    df = df.sort_values("Date").reset_index(drop=True)
    return df


def _save(df, name):
    path = OUTPUT_DIR / f"{name}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    nulls = df.isnull().sum()
    null_info = ", ".join(
        f"{c}:{n}" for c, n in nulls.items() if n > 0 and c != "Date"
    )
    print(f"  [{name}]  {len(df)} 行 x {len(df.columns)} 列  "
          f"| 日期 {df['Date'].min().date()} ~ {df['Date'].max().date()}")
    if null_info:
        print(f"    缺失值: {null_info}")
    print(f"    -> {path}")


def run():
    print(f"读取: {SOURCE_FILE}\n")

    # data_only=True 读取 Excel 保存后的缓存值
    wb = load_workbook(SOURCE_FILE, read_only=False, data_only=True)

    df_overseas = extract_overseas(wb)
    df_domestic = extract_domestic(wb)
    df_cnh      = extract_cnh(wb)
    wb.close()

    df_merged = merge_all(df_overseas, df_domestic, df_cnh)
    df_merged = add_derived_columns(df_merged)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("输出文件:")
    _save(df_overseas, "外盘下午三点价格")
    _save(df_domestic, "内盘收盘价")
    _save(df_cnh,      "离岸人民币")
    _save(df_merged,   "合并数据")
    print("\n完成。")

    return df_overseas, df_domestic, df_cnh, df_merged


if __name__ == "__main__":
    run()
